# Fabric notebook source

# METADATA ********************

# META {
# META   "kernel_info": {
# META     "name": "synapse_pyspark"
# META   },
# META   "dependencies": {
# META     "lakehouse": {
# META       "default_lakehouse": "b6adb647-6f2d-4cc8-b9e2-9a1201642b6a",
# META       "default_lakehouse_name": "den_lhw_dpr_001_policy_product",
# META       "default_lakehouse_workspace_id": "a5b83bde-449c-4623-a821-90f37a02ac15"
# META     },
# META     "environment": {
# META       "environmentId": "eccb61a4-306f-40f8-a7e1-53e1b34b5b1a",
# META       "workspaceId": "00000000-0000-0000-0000-000000000000"
# META     }
# META   }
# META }

# CELL ********************

from datetime import datetime, timedelta
import pandas as pd
from pyspark.sql import functions as F
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import datetime as dt
from spark_engine.common.email_util import send_email

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

%run den_nbk_pdi_001_workspace_parameters

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

%run den_nbk_pde_001_shared_utils

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# PARAMETERS CELL ********************

run_id = "00000000-0000-0000-0000-000000000000"
workspace_id = "00000000-0000-0000-0000-000000000000"
pipeline_name = "dfa_pln_dpr_001_gl_post_bind"
trigger_time = "1900-01-01T00:00:00Z"

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

today = dt.datetime.utcnow()
last_week = today - dt.timedelta(days=7)

run_date = today.strftime("%Y-%m-%d")

replacement_tokens = {
    "workspace_name": "Fabric Workspace",
    "run_date": run_date
}

email_template_path = get_template_location_url(
    file_name="gl_ln_naics_response.json"
)

email_json = read_json_file(email_template_path)
email_dict = replace_tokens_in_json_object(email_json, replacement_tokens)

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

gl_df = (
    spark.table("policy.fact_gl_cdpf_enrichment")
    .filter(F.col("request_ts") >= F.lit(last_week))
)

pdf = gl_df.toPandas()

# The enrichment table is append-only: every post-bind run INSERTs a fresh copy
# of its rows (see den_nbk_pd_001_gl_post_bind). So the 7-day window can contain
# (a) exact duplicates if the post-bind ran more than once, and (b) stale rows
# from an earlier code version (e.g. blank payroll / exposure basis, old "Y"
# blacklist flag). Keep only the most recent run per (policy_code, class_codes)
# so the workbook shows one clean, current row -- the latest run is the one with
# payroll/exposure populated, so latest-wins also drops the stale blank rows.
if not pdf.empty and "request_ts" in pdf.columns:
    _dedup_key = [c for c in ["policy_code", "class_codes"] if c in pdf.columns]
    if _dedup_key:
        pdf = (
            pdf.sort_values("request_ts")
               .drop_duplicates(subset=_dedup_key, keep="last")
               .reset_index(drop=True)
        )

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

FINAL_COLUMNS = [
    "TYPE OF CONCERN",
    "FINAL AUDIT RESULTS",
    "Audit Result NAICS Code",
    "Condition Score",
    "Endt Amount",
    "Agency Feedback Form Needed?",
    "Notes - Brief Summary of Concern(s)",
    "DAYS TO/PAST INCEPTION",
    "AUDIT DATE",
    "POST BIND UW",

    "Lexis Nexis Confidence Score",
    "NAICS Code (Lexis Nexis)",
    "NAICS Code 2 (Lexis Nexis)",
    "NAICS Code 3 (Lexis Nexis)",

    "NAICS Blacklist Flag",
    "LN Response Flag",

    "NAICS Code (Agent Entered)",
    "Policy Code",
    "Insured Name",
    "DBA (Doing Business As)",
    "Address",
    "Class Codes (Agent entered)",
    "Description of Class Codes",
    "Suffixes",
    "Agency Code",
    "Inception",
    "Agency UW",
    "YTD Prem",
    "Gov State",
    "Industry",
    "Sub Industry",
    "Bus. Type",
    "Exposure Basis",
    "Payroll (Exposure)",
    "Revenue (Exposure)",
]

def map_row(r):
    # Build combined address
    parts = [
        str(r.get("street_number") or "").strip(),
        str(r.get("street_name") or "").strip(),
        str(r.get("city") or "").strip(),
        str(r.get("state") or "").strip(),
        str(r.get("zip5") or "").strip(),
    ]
    address = ", ".join(p for p in parts if p)

    return {
        **{col: "" for col in FINAL_COLUMNS[0:10]},
        "Lexis Nexis Confidence Score": r.get("confidence_score"),
        "NAICS Code (Lexis Nexis)": r.get("naics_code"),
        "NAICS Code 2 (Lexis Nexis)": r.get("naics_2"),
        "NAICS Code 3 (Lexis Nexis)": r.get("naics_3"),
        "NAICS Blacklist Flag": r.get("naics_blacklist_flag"),
        "LN Response Flag": r.get("ln_response_flag"),

        "NAICS Code (Agent Entered)": r.get("naics_code_agent"),
        "Policy Code": r.get("policy_code"),
        "Insured Name": r.get("insured_name"),
        "DBA (Doing Business As)": r.get("dba"),
        "Address": address,
        "Class Codes (Agent entered)": r.get("class_codes"),
        "Description of Class Codes": r.get("class_desc"),
        "Suffixes": r.get("suffixes"),
        "Agency Code": r.get("agency_code"),
        "Inception": r.get("inception"),
        "Agency UW": r.get("agency_uw"),
        "YTD Prem": r.get("ytd_prem"),
        "Gov State": r.get("state"),
        "Industry": r.get("industry"),
        "Sub Industry": r.get("sub_industry"),
        "Bus. Type": r.get("bus_type"),
        "Exposure Basis": r.get("exposure_basis"),
        "Payroll (Exposure)": r.get("payroll"),
        "Revenue (Exposure)": r.get("revenue"),
    }

final_df = pd.DataFrame([map_row(r) for _, r in pdf.iterrows()], columns=FINAL_COLUMNS)

file_name = f"gl_policy_ln_weekly_{run_date}.xlsx"
local_file_path = f"/lakehouse/default/Files/{file_name}"
final_df.to_excel(local_file_path, index=False)

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

wb = load_workbook(local_file_path)
ws = wb.active

orange = PatternFill(fill_type="solid", fgColor="F4B183")
yellow = PatternFill(fill_type="solid", fgColor="FFF2CC")
blue = PatternFill(fill_type="solid", fgColor="D9EAF7")
bold = Font(bold=True)

for idx, cell in enumerate(ws[1], start=1):
    if idx <= 10:
        cell.fill = orange
    elif idx <= 12:
        cell.fill = yellow
    elif idx <= 14:
        cell.fill = blue

    cell.font = bold
    cell.alignment = Alignment(wrap_text=True)

wb.save(local_file_path)

attachment_path = get_file_location_url(
    lakehouse_name="den_lhw_dpr_001_policy_product",
    file_relative_path=file_name
)

print("âœ… Excel styled")

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

input_params = {
    "subject": email_dict.get("subject"),
    "body": email_dict.get("body", {}).get("content", ""),
    "to_email": email_dict.get("emailRecipient"),
    "cc_email": email_dict.get("emailCc"),
    "from_account": email_dict.get("emailSender"),
    "key_vault_name": secretsScope,
    "attachments": [attachment_path],
}

send_email(**input_params)
print("✅ Email sent")

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }
